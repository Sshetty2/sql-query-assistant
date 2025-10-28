"""Excel I/O utilities for FK inferencing audit trail."""

import openpyxl
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Dict, Optional

from database.infer_foreign_keys import has_existing_fk


def create_excel(excel_path: str, id_columns: List[Tuple], existing_fks: Dict):
    """
    Create Excel with pre-populated ID columns.

    Args:
        excel_path: Path to Excel file
        id_columns: List of (table_name, column_name, base_name) tuples
        existing_fks: Dict mapping table_name to list of existing FKs
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FK_Mappings"

    # Headers
    headers = [
        "table_name", "fk_column", "base_name",
        "candidate_1", "score_1", "candidate_2", "score_2",
        "candidate_3", "score_3", "candidate_4", "score_4",
        "candidate_5", "score_5",
        "chosen_table", "chosen_score", "decision_type", "timestamp", "notes"
    ]
    ws.append(headers)

    # Pre-populate rows
    for table_name, column_name, base_name in id_columns:
        if has_existing_fk(column_name, existing_fks.get(table_name, [])):
            ws.append([
                table_name, column_name, base_name,
                "[EXISTING]", "N/A", "", "", "", "", "", "", "", "",
                "[EXISTING]", "N/A", "existing", "", "Explicit FK exists"
            ])
        else:
            ws.append([table_name, column_name, base_name] + [""] * 14)

    wb.save(excel_path)
    print(f"[PASS] Created Excel: {excel_path} ({len(id_columns)} ID columns)")


def find_next_incomplete_row(excel_path: str) -> Optional[int]:
    """
    Find first row where chosen_table is empty.

    Args:
        excel_path: Path to Excel file

    Returns:
        Row index (1-based) or None if all complete
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["FK_Mappings"]

    for row_idx in range(2, ws.max_row + 1):  # Skip header
        if not ws.cell(row_idx, 14).value:  # Column N: chosen_table
            return row_idx
    return None


def load_row_data(excel_path: str, row_idx: int) -> Dict:
    """
    Load data from a specific row.

    Args:
        excel_path: Path to Excel file
        row_idx: Row index (1-based)

    Returns:
        Dict with table_name, fk_column, base_name
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["FK_Mappings"]

    return {
        "table_name": ws.cell(row_idx, 1).value,
        "fk_column": ws.cell(row_idx, 2).value,
        "base_name": ws.cell(row_idx, 3).value
    }


def write_candidates(excel_path: str, row_idx: int, candidates: List[Tuple]):
    """
    Write candidates to Excel row.

    Args:
        excel_path: Path to Excel file
        row_idx: Row index (1-based)
        candidates: List of (table_name, score) tuples
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["FK_Mappings"]

    for i, (table, score) in enumerate(candidates[:5]):
        ws.cell(row_idx, 4 + i*2).value = table
        ws.cell(row_idx, 5 + i*2).value = round(score, 3)

    wb.save(excel_path)


def write_decision(excel_path: str, row_idx: int, decision: Dict):
    """
    Write decision to Excel row.

    Args:
        excel_path: Path to Excel file
        row_idx: Row index (1-based)
        decision: Dict with chosen_table, chosen_score, decision_type, notes
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["FK_Mappings"]

    ws.cell(row_idx, 14).value = decision["chosen_table"]
    ws.cell(row_idx, 15).value = decision.get("chosen_score", "")
    ws.cell(row_idx, 16).value = decision["decision_type"]
    ws.cell(row_idx, 17).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row_idx, 18).value = decision["notes"]

    wb.save(excel_path)


def get_statistics(excel_path: str) -> Dict:
    """
    Calculate statistics from completed Excel file.

    Args:
        excel_path: Path to Excel file

    Returns:
        Dict with counts by decision_type
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["FK_Mappings"]

    stats = {
        "total": ws.max_row - 1,  # Exclude header
        "auto": 0,
        "manual": 0,
        "existing": 0,
        "skipped": 0,
        "incomplete": 0
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        decision_type = row[15].value  # Column P: decision_type (0-indexed: 15 = column 16)
        if decision_type:
            stats[decision_type] = stats.get(decision_type, 0) + 1
        else:
            stats["incomplete"] += 1

    return stats
